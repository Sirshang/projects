''''''
'''- 读取Excel文件中数据的步骤：
  - 1、打开excel文件
  - 2、获取sheet
  - 3、读取操作
  - 4、写入操作，一定要记得保存文件
'''
import openpyxl

# file = 'D:\Pycharm_workspace\py61\d13_0602_python单元测试框架pytest_2\测试数据.xlsx'
# openpyxl.load_workbook(filename = file)
book = openpyxl.load_workbook(filename = '测试数据.xlsx')
sh1 = book['登录']
# 读取某一个单元格数据
# excepted1 = sh1.cell(2,3).value
# print(excepted1)

# 读取所有的数据：[(),(),()]
all_data = list(sh1.values)
print(all_data)  # 生成器

# 写入数据
# sh1.cell(6,1).value = '5'
# book.save(filename = '测试数据.xlsx')

# 针对读取的数据进行格式转换 :  [{},{},{}]
new_all_data = []
title = all_data[0]  # ('用例编号', '用例标题', '用户名', '密码', '期望结果')
for case in all_data[1:]:
    new_case = dict(zip(title,case))
    new_all_data.append(new_case)
print(new_all_data)

# excel的函数封装：快去的去读取任意的.xlsx下的任意sheet中的数据
def read_excel(file_name,sheet_name):
    '''
    读取excel中某个sheet所有的数据
    :param file_name: excel文件的路径和名称，格式.xlsx
    :param sheet_name: excel文件中sheet的名称
    :return: 所有的数据，列表嵌套字典的格式
    '''
    book = openpyxl.load_workbook(filename=file_name)
    sh1 = book[sheet_name]
    all_data = list(sh1.values)
    new_all_data = []
    title = all_data[0]
    for case in all_data[1:]:
        new_case = dict(zip(title, case))
        new_all_data.append(new_case)
    return new_all_data


